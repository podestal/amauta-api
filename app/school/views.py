from datetime import date
from unidecode import unidecode
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils.timezone import now
from django.db.models.functions import ExtractWeek, ExtractMonth, ExtractDay
from django.utils import timezone
from datetime import timedelta
from rest_framework.viewsets import ModelViewSet
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAdminUser, SAFE_METHODS, IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from rest_framework import status
from django.db.models import Subquery, OuterRef, Prefetch
from django.core.cache import cache
from datetime import datetime

import pandas as pd
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from notification.push_notifications import send_push_notification
from . import tasks
from notification.models import FCMDevice

from django.http import JsonResponse
from twilio.rest import Client
from django.conf import settings

from . import models
from . import serializers

class CustomPagination(PageNumberPagination):
    page_size = 10  
    page_size_query_param = 'page_size' 
    max_page_size = 100  

class AreaViewSet(ModelViewSet):
    queryset = models.Area.objects.all()
    serializer_class = serializers.AreaSerializer
    
    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [IsAuthenticated()]
        return [IsAdminUser()]

class SchoolViewSet(ModelViewSet):
    queryset = models.School.objects.all()
    serializer_class = serializers.SchoolSerializer
    permission_classes = [IsAuthenticated]
    
    # def get_permissions(self):
    #     if self.request.method in SAFE_METHODS:
    #         return [IsAuthenticated()]
    #     return [IsAdminUser()]

class CompetenceViewSet(ModelViewSet):
    queryset = models.Competence.objects.select_related('area')
    serializer_class = serializers.CompetenceSerializer

    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [IsAuthenticated()]
        return [IsAdminUser()]

class CapacityViewSet(ModelViewSet):
    queryset = models.Capacity.objects.select_related('competence')
    serializer_class = serializers.CapacitySerializer

    def get_permissions(self):
        if self.request.method in SAFE_METHODS:
            return [IsAuthenticated()]
        return [IsAdminUser()]

class ClaseViewSet(ModelViewSet):

    queryset = models.Clase.objects.select_related('school').prefetch_related('students')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateClaseSerializer
        if self.request.method == 'DELETE':
            return serializers.RemoveClaseSerializer
        return serializers.GetClaseSerializer
    
    # def get_permissions(self):
    #     if self.request.method in SAFE_METHODS:
    #         return [IsAuthenticated()]
    #     return [IsAdminUser()]
    
    def get_queryset(self):
        if self.request.user.is_superuser:
            return super().get_queryset()
        school = self.request.query_params.get('school')
        if not school:
            return models.Clase.objects.none()
        return self.queryset.filter(school=school)
    
    # @action(detail=False, methods=['get'])
    # def withStudentsCount(self, request):

    #     school = request.query_params.get('school')
    #     if not school:
    #         return Response({"error": "School parameter is required"}, status=400)
    #     clases = self.queryset.filter(school=school)
    #     serializer = serializers.GetClaseForSummarySerializer(clases, many=True)
    #     return Response(serializer.data)
        

class InstructorViewSet(ModelViewSet):

    queryset = models.Instructor.objects.select_related('user', 'school').prefetch_related('clases')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateInstructorSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return serializers.UpdateInstructorSerializer
        return serializers.GetInstructorSerializer
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        user = self.request.user
        try:
            instructor = self.queryset.get(user=user)
        except:
            raise NotFound("Instructor not found for the current user.")
        serializer = serializers.GetInstructorSerializer(instructor)
        return Response(serializer.data)
    
class ManagerViewSet(ModelViewSet):

    queryset = models.Manager.objects.select_related('user', 'school')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateManagerSerializer
        return serializers.GetManagerSerializer
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        user = self.request.user
        try:
            manager = self.queryset.get(user=user)
        except:
            raise NotFound("Manager not found for the current user.")
        serializer = serializers.GetManagerSerializer(manager)
        return Response(serializer.data)
    
class AtendanceViewSet(ModelViewSet):

    queryset = models.Atendance.objects.select_related('student')
    # permission_classes = [IsAuthenticated]
    # permission_classes = [IsAdminUser]

        
    def get_permissions(self):
        # print('user', self.request.user.groups.all()[0].name)
        if self.request.user.groups.all()[0].name == 'instructor':
            instructor = models.Instructor.objects.get(user=self.request.user)
            school = models.School.objects.get(id=instructor.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]
        
        if self.request.user.groups.all()[0].name == 'manager':
            manager = models.Manager.objects.get(user=self.request.user)
            school = models.School.objects.get(id=manager.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]
        
        if self.request.user.groups.all()[0].name == 'assistant':
            assistant = models.Assistant.objects.get(user=self.request.user)
            school = models.School.objects.get(id=assistant.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]

        return [IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateAtendanceSerializer
        return serializers.GetAtendanceSerializer
    
    # def get_permissions(self):
    #     if self.request.method == 'DELETE':
    #         return [IsAdminUser()]
    #     return [IsAuthenticated()]

    @action(detail=False, methods=['get'])
    def byClassroom(self, request):
        '''Get all attendances by classroom and dates'''
        attendances = self.queryset
        classroom = request.query_params.get('classroom')
        week_param = request.query_params.get('week')
        day_param = request.query_params.get('day')
        month_param = request.query_params.get('month')

        if not classroom:
            return Response({"error": "Classroom parameter is required"}, status=400)
        if (week_param):
            attendances = attendances.filter(student__clase=classroom, created_at__week=week_param )
        if (day_param):
            attendances = attendances.filter(student__clase=classroom, created_at__day=day_param, created_at__month=month_param)
        if (month_param):
            attendances = attendances.filter(student__clase=classroom, created_at__month=month_param)
        serializer = serializers.GetSimpleAttendanceSerializer(attendances, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byStudent(self, request):
        student_id = request.query_params.get('student')
        if not student_id:
            return Response({"error": "Student parameter is required"}, status=400)
        month = request.query_params.get('month')
        if not month:
            month = datetime.today().month
        student = get_object_or_404(models.Student, uid=student_id)
        attendances = self.queryset.filter(student=student, created_at__month=month).order_by('-created_at')
        if not attendances.exists():
            return Response([])
        serializer = serializers.GetAtendanceSerializer(attendances, many=True)
        return Response(serializer.data)
    
    def save_to_cache(self, student, kind, status, request, attendance_id=None):

        cache_key = f"attendance_{student.uid}_{kind}"
        cache_data = {
            "id": attendance_id,
            "status": status,
            "kind": kind,
            "created_by": request.data.get('created_by'),
            "observations": request.data.get('observations', ''),
            "created_at": datetime.now().isoformat(),
        }

        cache.set(cache_key, cache_data, timeout=64800)
        print('Cache created', cache_key)
        return cache_data
    
    def get_notification_message(self, student, status):

        message = ''

        if status == 'L':
            message = f'{student.first_name} llegó tarde'
        elif status == 'N':
            message = f'{student.first_name} no asistió'
        elif status == 'T':
            message = f'{student.first_name} se retiró temprano'
        elif status == 'E':
            message = f'{student.first_name} fué excusado'

        return message

    def update(self, request, *args, **kwargs):
        print('request', request.data)

        status = request.data['status']
        kind = request.data['kind']
        student_id = request.data['student']
        student = models.Student.objects.get(uid=student_id)
    
        if not student: 
            return Response({"error": "No se pudo encontrar alumno"}, status=400)
        notification_message = self.get_notification_message(student, status)
        try:
            users = models.Tutor.objects.filter(students__uid=student_id).values_list('user', flat=True)
            # def send_attendance_notification(users, notification_message, apologize_message=None):
        except:
            print('could not find users')
        tasks.send_attendance_notification.delay(list(users), notification_message, apologize_message=f'{student.first_name} llegó a tiempo')
        return super().update(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        print('data', request.data)
        status = request.data['status']
        kind = request.data['kind']
        student_id = request.data['student']
        attendance_type = request.data['attendance_type']

        try:
            student = models.Student.objects.get(uid=student_id)
        except:
            return Response({"error": "No se pudo encontrar alumno"}, status=400)

        existing_attendance = models.Atendance.objects.filter(
            student__uid=student_id,
            created_at__date=date.today()
        )

        if existing_attendance.count() == 2:
            return Response({"error": "Alumno ya fué escaneado"}, status=400)
        
        if existing_attendance.count() == 1 and kind == 'I':
            return Response({"error": "Alumno ya fué escaneado"}, status=400)
        
        if existing_attendance.count() == 0 and kind == 'O':
            return Response({"error": "Alumno no asistió"}, status=400)

        now = datetime.now().time()

        if attendance_type == 'A' and kind == 'I' and status != 'E' and status != 'T':
            status = 'O' if now < student.school.automatic_late else 'L'
        
        request.data['status'] = status

        if status != 'O':
            notification_message = self.get_notification_message(student, status)
            try:
                users = models.Tutor.objects.filter(students__uid=student_id).values_list('user', flat=True)
            except:
                return super().create(request, *args, **kwargs)
            
            tasks.send_attendance_notification.delay(list(users), notification_message)


        if status == 'N':
            no_assist_announcement = models.Announcement.objects.create(
                title='Falta',
                description=f'{student.first_name} {student.last_name}  no asistió a clases',
                announcement_type='E',
                visibility_level='P',
                school=student.school,
                created_by=request.user if request.user.is_authenticated else None
            )

            no_assist_announcement.students.set([student])
        
        if status == 'L':
            late_announcement = models.Announcement.objects.create(
                title='Tardanza',
                description=f'{student.first_name} {student.last_name} llegó tarde a clases',
                announcement_type='A',
                visibility_level='P',
                school=student.school,
                created_by=request.user if request.user.is_authenticated else None
            )

            late_announcement.students.set([student])
 
        return super().create(request, *args, **kwargs)
    
class AssistantViewSet(ModelViewSet):

    queryset = models.Assistant.objects.select_related('user', 'school').prefetch_related('clases')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateAssistantSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return serializers.UpdateAssistantSerializer
        return serializers.GetAssistantSerializer
    
    # def get_permissions(self):
    #     if self.request.method in SAFE_METHODS or self.request.method in ['PUT', 'PATCH']:
    #         return [IsAuthenticated()]
    #     return [IsAdminUser()]
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        user = self.request.user
        try:
            assistant = self.queryset.get(user=user)
        except:
            return Response({"error": "Assistant not found for the current user"}, status=404)
        serializer = serializers.GetAssistantSerializer(assistant)
        return Response(serializer.data)

class StudentViewSet(ModelViewSet):

    # permission_classes = [IsAuthenticated]
    # permission_classes = [IsAdminUser]

    def get_permissions(self):
        # print('user', self.request.user.groups.all()[0].name)
        if self.request.user.groups.all()[0].name == 'instructor':
            instructor = models.Instructor.objects.get(user=self.request.user)
            school = models.School.objects.get(id=instructor.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]
        
        if self.request.user.groups.all()[0].name == 'manager':
            manager = models.Manager.objects.get(user=self.request.user)
            school = models.School.objects.get(id=manager.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]
        
        if self.request.user.groups.all()[0].name == 'assistant':
            assistant = models.Assistant.objects.get(user=self.request.user)
            school = models.School.objects.get(id=assistant.school.id)
            # if school.payment_status == 'P':
            #     return [IsAuthenticated()]
            # else:
            #     return [IsAdminUser()]

        return [IsAuthenticated()]

    def get_queryset(self):

        now = timezone.now()
        current_month = now.month

        day_param = self.request.query_params.get('day')
        week_param = self.request.query_params.get('week')
        month_param = self.request.query_params.get('month', current_month)

        attendance_in = models.Atendance.objects.filter(
            kind='I'
        )
        attendance_out = models.Atendance.objects.filter(
            kind='O'
        )

        if day_param:
            attendance_in = attendance_in.annotate(day=ExtractDay('created_at'), month=ExtractMonth('created_at')).filter(day=int(day_param), month=int(month_param))
            attendance_out = attendance_out.annotate(day=ExtractDay('created_at'), month=ExtractMonth('created_at')).filter(day=int(day_param), month=int(month_param))

        elif week_param:
            attendance_in = attendance_in.annotate(week=ExtractWeek('created_at')).filter(week=int(week_param))
            attendance_out = attendance_out.annotate(week=ExtractWeek('created_at')).filter(week=int(week_param))
        
        elif month_param:
            attendance_in = attendance_in.annotate(month=ExtractMonth('created_at')).filter(month=int(month_param))
            attendance_out = attendance_out.annotate(month=ExtractMonth('created_at')).filter(month=int(month_param))

        else:
            attendance_in = attendance_in.filter(created_at=now.today())
            attendance_out = attendance_out.filter(created_at=now.today())

        return (
                models.Student.objects.select_related('clase', 'school').prefetch_related('health_info', 'birth_info', 'emergency_contact', 'tutors', 'averages', 'read_agendas', 'tutor_contact')
                .prefetch_related(
                    Prefetch('atendances', queryset=attendance_in, to_attr='attendance_in'),
                    Prefetch('atendances', queryset=attendance_out, to_attr='attendance_out'),
                )
            )

    def get_serializer_class(self):

        # if self.action == 'list':
        #     return serializers.GetStudentsSerializer
        if self.request.method == 'POST':
            return serializers.CreateStudentSerializer
        if self.request.method in ['PUT', 'PATCH']:
            return serializers.UpdateStudentSerializer
        return serializers.GetStudentSerializer

    @action(detail=False, methods=['get'])
    def byClassroom(self, request):
        classroom = request.query_params.get('classroom')
        if not classroom:
            return Response({"error": "Classroom parameter is required"}, status=400)
        
        try:
            classroom = int(classroom)
        except ValueError:
            return Response({"error": "Classroom parameter must be an integer"}, status=400)

        students = self.get_queryset().filter(clase=classroom)
        if not students.exists():
            return Response([], status=200)

        serializer = serializers.GetStudentSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byTutor(self, request):

        user_id = request.user.id
        try: 
            tutor = models.Tutor.objects.get(user_id=user_id)
            students = tutor.students.all()
            serializer = serializers.GetStudentForTutorSerializer(students, many=True)
            return Response(serializer.data)
        except:
            return Response({"error": "Tutor not found"}, status=404)

    @action(detail=False, methods=['get'])
    def byDni(self, request):
        dni = request.query_params.get('dni')
        if not dni:
            return Response({"error": "DNI parameter is required"}, status=400)
        student = get_object_or_404(models.Student, dni=dni)
        serializer = serializers.GetStudentSerializer(student)
        return Response(serializer.data)
        
    @action(detail=False, methods=['get'])
    def byName(self, request):
        school = request.query_params.get('school')
        name = request.query_params.get('name')

        if not name:
            return Response({"error": "Name parameter is required"}, status=400)

        normalized_name = unidecode(name)  # Removes accents

        query = (
            Q(first_name__icontains=normalized_name) | 
            Q(last_name__icontains=normalized_name)
        ) & Q(school_id=school)

        students = self.get_queryset().filter(query)
        serializer = serializers.GetStudentSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byAgendas(self, request):
        school = request.query_params.get('school')
        classroom = request.query_params.get('classroom')
        students = self.get_queryset().filter(school=school, clase=classroom)
        students = students.prefetch_related(
            Prefetch(
                'read_agendas',
                queryset=models.TutorReadAgenda.objects.filter(
                    agenda_date=timezone.now().date()
                ),
                to_attr='filtered_read_agendas'
            ),
            Prefetch(
                'tutor_contact',
                queryset=models.TutorContact.objects.filter(
                    contact_date=timezone.now().date()
                ),
                to_attr='filtered_tutor_contact'
            )
        )
        serializer = serializers.GetStudentByAgendaSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byLastTen(self, request):
        school = request.query_params.get('school')
        students = self.get_queryset().filter(school=school).order_by('-created_at')[:10]
        serializer = serializers.GetStudentSerializer(students, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def byTotalScore(self, request):
        '''Get all students with their total score'''
        classroom = request.query_params.get('classroom')   
        if not classroom:
            return Response({"error": "Classroom parameter is required"}, status=400)
        quarter = request.query_params.get('quarter')
        if not quarter:
            return Response({"error": "Quarter parameter is required"}, status=400)

        students = self.get_queryset().filter(clase=classroom)
        students = students.prefetch_related(
            Prefetch(
                'averages',
                queryset=models.QuarterGrade.objects.filter(
                    quarter=quarter
                ),
                to_attr='filtered_averages'
            )
        )
        print('students', students)
        serializer = serializers.GetStudentForTotalScoreSerializer(students, many=True)
        return Response(serializer.data)
    


    
    @action(detail=False, methods=['get'])
    def byQuarterGrade(self, request):
        clase = request.query_params.get('clase')
        competencies = request.query_params.get('competencies').split(',')
        competencies = [c for c in competencies if c]
        quarter = request.query_params.get('quarter')
        students = self.get_queryset().filter(clase=clase)
        students = students.prefetch_related(
            Prefetch(
                "averages",
                queryset=models.QuarterGrade.objects.filter(
                    quarter=quarter,
                    competence__in=competencies
                ),
                to_attr="filtered_averages"
            )
        )
        serializer = serializers.GetStudentForQuarterGradeSerializer(students, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byGrade(self, request):
        clase = request.query_params.get('clase')
        competence = request.query_params.get('competence')
        quarter = request.query_params.get('quarter')
        students = self.get_queryset().filter(clase=clase)
        students = students.prefetch_related(
            Prefetch(
                "grades",
                queryset=models.Grade.objects.filter(
                    activity__competences=competence,
                    activity__quarter=quarter
                )
                .select_related('activity', 'activity__category', 'assignature')
                .prefetch_related('activity__competences'),
                to_attr="filtered_grades"
            ),
            Prefetch(
                "averages",
                queryset=models.QuarterGrade.objects.filter(
                    quarter=quarter
                ),
                to_attr="filtered_averages"
            )
        )

        serializer = serializers.GetStudentForFilteredGradesSerializer(students, many=True)
        return Response(serializer.data)
    
    # @action(detail=False, methods=["get"])
    # def export_to_excel(self, request):
    #     # Fetch data from model
    #     data = list(models.Student.objects.values("uid", "first_name", "last_name"))  # Adjust fields
    #     df = pd.DataFrame(data)

    #     # Create an HTTP response
    #     response = HttpResponse(
    #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #     )
    #     response["Content-Disposition"] = 'attachment; filename="students.xlsx"'

    #     # Save the DataFrame to the response
    #     with pd.ExcelWriter(response, engine="openpyxl") as writer:
    #         df.to_excel(writer, sheet_name="Students", index=False)

    #     return response
    # @action(detail=False, methods=["get"])
    # def export_to_excel(self, request):
    #     # Create a new workbook and remove the default sheet
    #     wb = Workbook()
    #     wb.remove(wb.active)

    #     # Define data sources (Adjust fields as needed)
    #     student_data = list(models.Student.objects.values("uid", "first_name", "last_name"))
    #     instructor_data = list(models.Instructor.objects.values("id", "first_name", "last_name"))

    #     # Convert to DataFrame
    #     df_students = pd.DataFrame(student_data)
    #     df_instructors = pd.DataFrame(instructor_data)

    #     # Add sheets and write data
    #     self.write_to_sheet(wb, df_students, "Students")
    #     self.write_to_sheet(wb, df_instructors, "Instructors")

    #     # Prepare response
    #     response = HttpResponse(
    #         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    #     )
    #     response["Content-Disposition"] = 'attachment; filename="school_data.xlsx"'

    #     # Save workbook to response
    #     wb.save(response)
    #     return response

    # def write_to_sheet(self, wb, df, sheet_name):
    #     sheet = wb.create_sheet(title=sheet_name)

    #     # Write headers with styling
    #     header_font = Font(bold=True)
    #     for col_num, column_title in enumerate(df.columns, 1):
    #         cell = sheet.cell(row=1, column=col_num, value=column_title)
    #         cell.font = header_font
    #         cell.alignment = Alignment(horizontal="center")

    #     # Write data rows
    #     for row_num, row in enumerate(df.itertuples(index=False), 2):
    #         for col_num, value in enumerate(row, 1):
    #             sheet.cell(row=row_num, column=col_num, value=value)

    #     # Auto-adjust column width
    #     for col in sheet.columns:
    #         max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    #         sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    @action(detail=False, methods=["get"], url_path='export_info_to_excel', url_name='export_info_to_excel')
    def export_info_to_excel(self, request):
        school = request.query_params.get('school')
        students = self.get_queryset().filter(school=school)
        
        religion_choices = {
            'C': 'Católica',
            'E': 'Evangélica',
            'J': 'Judía',
            'M': 'Musulmana',
            'B': 'Budista',
            'O': 'Otra',
        }

        language_choices = {
            'S': 'Español',
            'E': 'Inglés',
            'Q': 'Quechua',
            'A': 'Aymara',
        }

        def get_grade_description(grade):
            grades_choices = {
                '1': '1ro',
                '2': '2do',
                '3': '3ro',
                '4': '4to',
                '5': '5to',
                '6': '6to',
            }
            return grades_choices.get(grade, grade)
        
        def get_section_description(section):
            if section == 'U':
                section = 'Unica'
            return f'{section}'
        
        def get_level_description(level):
            levels_choices = {
                'I': 'Inicial',
                'P': 'Primaria',
                'S': 'Secundaria',
            }
            return levels_choices.get(level, level)

        # Header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"

        # Define border style
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                            top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="000066", end_color="000066", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        # Define headers
        headers = [
            "DNI",
            "Nombres",
            "Apellidos",
            "Grado",
            "Sección",
            "Nivel",
            "Religión",
            "Lengua materna",
            "Segunda lengua",
            "Teléfono del Apoderado",
        ]
        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Write data to the sheet
        for student in students:
            row = [
                student.dni,
                student.first_name,
                student.last_name,
                get_grade_description(student.clase.grade),
                get_section_description(student.clase.section),
                get_level_description(student.clase.level),
                religion_choices.get(student.religion, '-'),
                language_choices.get(student.main_language, '-'),
                language_choices.get(student.second_language, '-'),
                student.tutor_phone
            ]
            ws.append(row)

        for col in ws.columns:
            max_length = 0
            column = col[0].column  
            column_letter = get_column_letter(column)  

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = max_length + 2  # Add a little padding
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        
        return response
        


    @action(detail=False, methods=["get"])
    def export_to_excel(self, request):

        quarter_description = {
            'Q1': 'Primer Bimestre',
            'Q2': 'Segundo Bimestre',
            'Q3': 'Tercer Bimestre',
            'Q4': 'Cuarto Bimestre',
        }


        classroom_grade_converter = {
            '1': 'Primero',
            '2': 'Segundo',
            '3': 'Tercero',
            '4': 'Cuarto',
            '5': 'Quinto',
            '6': 'Sexto',
        }

        classroom_level_converter = {
            'I': 'Inicial',
            'P': 'Primaria',
            'S': 'Secundaria',
        }


        # Create a workbook and rename the default sheet
        classroom = request.query_params.get('classroom')
        quarter= request.query_params.get('quarter')
        instructor_id = request.query_params.get('instructor_id')

        clase = get_object_or_404(models.Clase, id=classroom)
        classroom_section = clase.section if clase.section else ''
        if classroom_section == 'U':
            classroom_section = 'Única'
        classroom_grade = clase.grade if clase.grade else ''
        classroom_level = clase.level if clase.level else ''

        wb = openpyxl.Workbook()
        ws_general = wb.active
        ws_general.title = "Generalidades"

        # Define border style
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                            top=Side(style="thin"), bottom=Side(style="thin"))

        header_fill = PatternFill(start_color="000066", end_color="000066", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Format "DATOS GENERALES"
        ws_general["B2"] = "DATOS GENERALES :"
        ws_general["B2"].font = Font(bold=True, size=12)
        
        # "Institución Educativa" from B to J
        ws_general["B4"] = "Institución Educativa :"
        ws_general.merge_cells("B4:J4")
        ws_general["B4"].fill = header_fill
        ws_general["B4"].font = header_font
        ws_general["B4"].alignment = Alignment(horizontal="center")

        # "Código Modular - Anexo" from B to D
        ws_general["B5"] = "Código Modular - Anexo :"
        ws_general.merge_cells("B5:D5")
        ws_general.merge_cells("E5:F5")

        # "Nivel" at G, with space from H to J
        ws_general["G5"] = "Nivel :"
        ws_general['H5'] = f'{classroom_level_converter[classroom_level]}'
        ws_general.merge_cells("H5:J5")

        # "Nombre" at B, space from C to J
        ws_general["B6"] = "Nombre :"
        ws_general.merge_cells("B6:J6")

        # "Datos referentes al Registro de Notas" from B to J
        ws_general["B7"] = "Datos referentes al Registro de Notas :"
        ws_general.merge_cells("B7:J7")
        ws_general["B7"].fill = header_fill
        ws_general["B7"].font = header_font
        ws_general["B7"].alignment = Alignment(horizontal="center")

        # "Año Académico" from B to C, "2019" from D to J
        ws_general["B8"] = "Año Académico :"
        ws_general.merge_cells("B8:C8")
        ws_general["D8"] = "2025"
        ws_general.merge_cells("D8:J8")

        # "Diseño Curricular" from B to C, value from D to J
        ws_general["B9"] = "Diseño Curricular :"
        ws_general.merge_cells("B9:C9")
        ws_general["D9"] = "CURRÍCULO NACIONAL 2017"
        ws_general.merge_cells("D9:J9")

        # "Grado" from B to C, space from D to E, "Sección" at F, space from G to J
        ws_general["B10"] = "Periodo de evaluación :"
        ws_general.merge_cells("B10:C10")
        ws_general["D10"] = f"{quarter_description[quarter]}"
        ws_general.merge_cells("D10:F10")
        ws_general["G10"] = "Grado"
        ws_general["H10"] = f'{classroom_grade_converter[classroom_grade]}'
        ws_general["I10"] = "Sección"
        ws_general["J10"] = f'{classroom_section}'

        # Apply borders from B2 to J10
        for row in ws_general.iter_rows(min_row=4, max_row=10, min_col=2, max_col=10):
            for cell in row:
                cell.border = thin_border

        # "AREAS" Section
        ws_general["B13"] = "AREAS"
        ws_general["B13"].font = Font(bold=True)
        areas_ids = models.Assignature.objects.filter(clase=classroom,instructor=instructor_id).values_list('area_id', flat=True)
        # List areas from model in column B (ID) and column C (Title)
        areas = models.Area.objects.filter(id__in=areas_ids)
        row = 14
        for area in areas:
            ws_general[f"B{row}"] = area.id
            ws_general[f"C{row}"] = area.title
            row += 1

        competences = models.Competence.objects.select_related('area')
        students = self.get_queryset().filter(clase=classroom)  
        students = students.prefetch_related(
            Prefetch(
                "averages",
                queryset=models.QuarterGrade.objects.filter(
                    quarter='Q1',
                ),
                to_attr="filtered_averages"
            ))

        # Create separate sheets for each area
        for area in areas[:11]:  # Limit to 11 additional sheets
            ws_area = wb.create_sheet(title=area.title)
            ws_area["A1"] = 'ID'
            ws_area["A1"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws_area['A1'].font = Font(bold=True, color="FFFFFF")
            ws_area['A1'].alignment = Alignment(horizontal="center",  vertical="center")
            ws_area['B1'].font = Font(bold=True, color="FFFFFF")
            ws_area.merge_cells("A1:A2")

            ws_area["B1"] = 'Cod. Estudiante'
            ws_area["B1"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws_area["B1"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws_area['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws_area.merge_cells("B1:B2")

            ws_area["C1"] = 'Nombres'
            ws_area["C1"].font = Font(bold=True, color="FFFFFF")
            ws_area["C1"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            ws_area['C1'].alignment = Alignment(horizontal="center", vertical="center")
            ws_area.merge_cells("C1:C2")

            filtered_competences = competences.filter(area=area)

            start_col = 4 
            start_row = 1
            description_row = 2

            for index, competence in enumerate(filtered_competences, start=1):
                col_letter = get_column_letter(start_col)  # Get letter for merging
                next_col_letter = get_column_letter(start_col + 1)  # Next column for merging
                
                # Set the competence number
                ws_area[f"{col_letter}{start_row}"] = f"{index:02d}"  # Format 01, 02, etc.
                ws_area[f"{col_letter}{start_row}"].font = Font(bold=True, color="FFFFFF")
                ws_area[f"{col_letter}{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_area[f"{col_letter}{start_row}"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
                ws_area.merge_cells(f"{col_letter}{start_row}:{next_col_letter}{start_row}")  # Merge columns

                # Set the 'NL' label
                ws_area[f"{col_letter}{description_row}"] = "NL"
                ws_area[f"{col_letter}{description_row}"].font = Font(bold=True, color="FFFFFF")
                ws_area[f"{col_letter}{description_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_area[f"{col_letter}{description_row}"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

                # Set the description header
                ws_area[f"{next_col_letter}{description_row}"] = "Conclusión descriptiva de la competencia"
                ws_area[f"{next_col_letter}{description_row}"].font = Font(bold=True, color="FFFFFF")
                ws_area[f"{next_col_letter}{description_row}"].alignment = Alignment(horizontal="center", vertical="center")
                ws_area[f"{next_col_letter}{description_row}"].fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

                # Move to the next pair of columns
                start_col += 2
            
            start_row = 3
            for student in students:
                ws_area[f"A{start_row}"] = student.uid
                ws_area[f"B{start_row}"] = '0000004567'
                ws_area[f"C{start_row}"] = f"{student.first_name} {student.last_name}"
                    # Start placing grades and conclusions dynamically
                col = 4  # Start at first competence column
                for competence in filtered_competences:
                    col_letter = get_column_letter(col)
                    next_col_letter = get_column_letter(col + 1)

                    # Find the grade and conclusion for this competence
                    grade = ''
                    conclusion = ''
                    for avg in student.filtered_averages:
                        if avg.competence_id == competence.id:  # Match competence ID
                            grade = avg.calification
                            conclusion = avg.conclusion
                            break  # Stop once found
                    
                    # Insert the grade
                    ws_area[f"{col_letter}{start_row}"] = grade if grade else "NA"
                    ws_area[f"{col_letter}{start_row}"].alignment = Alignment(horizontal="left")

                    # Insert the conclusion
                    ws_area[f"{next_col_letter}{start_row}"] = conclusion if conclusion else ""
                    ws_area[f"{next_col_letter}{start_row}"].alignment = Alignment(horizontal="left", wrap_text=True)

                    col += 2  # Move to next competence column

                start_row += 1  # Move to the next student row


            ws_area[f"B{len(students) + 4}"] = 'LEYENDA'
            ws_area[f"B{len(students) + 4}"].font = Font(bold=True)

            ws_area[f"B{len(students) + 5}"] = 'NL= Nivel de logro alcanzado'
            ws_area.merge_cells(f"B{len(students) + 5}:C{len(students) + 5}")

            competence_start_row = len(students) + 6 

            for index, competence in enumerate(filtered_competences, start=1):
                row_number = competence_start_row + index - 1
                ws_area[f"B{row_number}"] = f"{index:02d}={competence.title}"  # Format: 01=Title
                ws_area.merge_cells(f"B{row_number}:C{row_number}")  # Merge B and C for better readability


            # Auto-adjust column width
            for col in ws_area.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                ws_area.column_dimensions[col_letter].width = max_length + 2

            for row in ws_area.iter_rows(min_row=1, max_row=start_row-1, min_col=1, max_col=3+(2*len(filtered_competences))):
                for cell in row:
                    cell.border = thin_border

        # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        
        return response
    

class TutorViewSet(ModelViewSet):

    queryset = models.Tutor.objects.select_related('user', 'school').prefetch_related('students', 'read_agendas')
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method in ['POST', 'PUT', 'PATCH']:
            return serializers.CreateTutorSerializer
        return serializers.GetTutorSerializer
    
    # def get_permissions(self):
    #     if self.request.method in SAFE_METHODS or self.request.method in ['PUT', 'PATCH']:
    #         return [IsAuthenticated()]
    #     return [IsAdminUser()]
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        user = self.request.user
        try:
            tutor = self.queryset.get(user=user)
        except:
            return Response({"error": "Tutor not found for the current user"}, status=404)
        serializer = serializers.GetTutorSerializer(tutor)
        return Response(serializer.data)

class CategoryViewSet(ModelViewSet):
    queryset = models.Category.objects.select_related('instructor')
    serializer_class = serializers.CategorySerializer  
    permission_classes = [IsAuthenticated]

    def get_queryset(self):

        if self.request.user.is_superuser:
            return super().get_queryset()
        user = self.request.user
        try:
            instructor = models.Instructor.objects.get(user_id=user.id)
        except:
            return Response({"error": "Instructor not found for the current user"}, status=404)
        return self.queryset.filter(instructor_id=instructor.id)
    
    def perform_destroy(self, instance):
        try:
            instance.delete()
        except:
            raise ValidationError("Esta categoría no puede ser eliminada porque tiene actividades asociadas.")
    
class AssignatureViewSet(ModelViewSet):
    queryset = models.Assignature.objects.select_related('clase', 'instructor', 'area')
    serializer_class = serializers.AssignatureSerializer  
    permission_classes = [IsAuthenticated]


    @action(detail=False, methods=['get'])
    def byInstructor(self, request):
        user = self.request.user
        try:
            instructor = models.Instructor.objects.get(user_id=user.id)
        except:
            return Response({"error": "Instructor not found for the current user"}, status=404)
        assignatures = self.queryset.filter(instructor_id=instructor.id)
        serializer = serializers.AssignatureSerializer(assignatures, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byTutor(self, request):
        studentUid = request.query_params.get('student')
        quarter = request.query_params.get('quarter')
        user = self.request.user
        try:
            tutor = models.Tutor.objects.get(user_id=user.id)
        except:
            return Response({"error": "Tutor not found for the current user"}, status=404)
        assignatures = self.queryset.filter(clase__students__tutors=tutor)
        serializer = serializers.GetAssignaturesForTutorSerializer(assignatures, many=True, context={'studentUid': studentUid, 'quarter': quarter})
        return Response(serializer.data)

class ActivityViewSet(ModelViewSet):

    queryset = models.Activity.objects.select_related('assignature', 'category').prefetch_related('competences', 'capacities', 'lessons')
    serializer_class = serializers.ActivitySerializer  
    # permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def byAssignature(self, request):
        assignature = request.query_params.get('assignature')
        competence = request.query_params.get('competence')
        quarter = request.query_params.get('quarter')
        if not assignature:
            return Response({"error": "Assignature parameter is required"}, status=400)
        activities = self.queryset.filter(assignature_id=assignature, quarter=quarter)
        if competence:
            activities = activities.filter(competences=competence)
        serializer = serializers.ActivitySerializer(activities, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def byTutor(self, request):
        assignature = request.query_params.get('assignature')
        studentUid = request.query_params.get('student')
        quarter = request.query_params.get('quarter')
        if not assignature:
            return Response({"error": "Assignature parameter is required"}, status=400)
        activities = self.queryset.filter(assignature_id=assignature, quarter=quarter)
        serializer = serializers.GetActivityForTutorSerializer(activities, many=True, context={'studentUid': studentUid, 'quarter': quarter})
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def byLesson(self, request):

        lessons = request.query_params.get('lessons').split(',')
        lessons = [l for l in lessons if l]
        if not lessons:
            return Response({"error": "Lesson parameter is required"}, status=400)
        activities = self.queryset.filter(lessons__in=lessons).distinct()
        if not activities.exists():
            return Response([], status=200)
        serializer = serializers.ActivitySerializer(activities, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):

        assignatureId = request.data['assignature']
        lessonsId = request.data['lessons'] if 'lessons' in request.data else ''
        assignature = models.Assignature.objects.get(id=assignatureId)
        students_uids = models.Student.objects.filter(clase=assignature.clase.id).values_list('uid', flat=True)
        users = models.Tutor.objects.filter(students__in=students_uids).values_list('user', flat=True)

        # creating the activity
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        activity = serializer.save()
        # setting up lessons to activity
        if lessonsId:
            activity.lessons.set(lessonsId)

        assignature_title = assignature.title
        activity_title = request.data['title']
        activity_due_date = request.data['due_date']
        activity_announcement = models.Announcement.objects.create(
            title='Nueva Actividad',
            description=f'{activity_title} del curso {assignature_title} para el día {activity_due_date}.',
            announcement_type='I',
            visibility_level='C',
            school=assignature.clase.school
        )

        activity_announcement.clases.set([assignature.clase])

        tasks.send_activity_notification.delay(list(users), activity_title, activity_due_date, 'Nueva Actividad', False)
        return Response(self.get_serializer(activity).data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        activity_title = request.data['title']
        activity_due_date = request.data['due_date']
        assignatureId = request.data['assignature']
        assignature = models.Assignature.objects.get(id=assignatureId)
        classroom = assignature.clase
        students_uids = models.Student.objects.filter(clase=classroom).values_list('uid', flat=True)
        users = models.Tutor.objects.filter(students__in=students_uids).values_list('user', flat=True)
        activity = super().update(request, *args, **kwargs)
        tasks.send_activity_notification.delay(list(users), activity_title, activity_due_date, 'Cambios en actividad', True)
        return activity
    
    

class GradeViewSet(ModelViewSet):

    queryset = models.Grade.objects.select_related('student', 'activity', 'assignature')
    serializer_class = serializers.GradeSerializer  
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return serializers.GradesByActivitySerializer
        return serializers.GradeSerializer  

    @action(detail=False, methods=['get'])
    def byActivity(self, request):
        activity = request.query_params.get('activity')
        if not activity:
            return Response({"error": "Activity parameter is required"}, status=400)
        grades = self.queryset.filter(activity_id=activity)
        serializer = serializers.GradesByActivitySerializer(grades, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byStudent(self, request):
        # activity quarter
        student_uid = request.query_params.get('student')
        if not student_uid:
            return Response({"error": "Student parameter is required"}, status=400)
        quarter = request.query_params.get('quarter')
        if not quarter:
            return Response({"error": "Quarter parameter is required"}, status=400)
        grades = self.queryset.filter(student__uid=student_uid, activity__quarter=quarter).select_related('activity__assignature')
        if not grades.exists():
            return Response([], status=200)
        serializer = serializers.GradesByStudentSerializer(grades, many=True)
        return Response(serializer.data)
    
    def update(self, request, *args, **kwargs):
        student_id = self.request.query_params.get('student_uid')
        student = models.Student.objects.get(uid=student_id)
        users = models.Tutor.objects.filter(students__uid=student_id).values_list('user', flat=True)
        updated_grade = super().update(request, *args, **kwargs)
        notification_message = f'{student.first_name} ha recibido una nueva calificación'
        tasks.send_grade_notification.delay(list(users), notification_message)

        grade = models.Grade.objects.get(id=updated_grade.data['id'])
        if updated_grade.data['calification'] == 'C':
            grade_announcement = models.Announcement.objects.create(
                title='Nueva Calificación',
                description=f'{student.first_name} {student.last_name} ha recibido una calificación muy baja en la actividad {grade.activity.title}.',
                announcement_type='E',
                visibility_level='P',
                school=student.school
            )
            grade_announcement.students.set([student])
        elif updated_grade.data['calification'] == 'B':
            grade_announcement = models.Announcement.objects.create(
                title='Nueva Calificación',
                description=f'{student.first_name} {student.last_name} ha recibido una calificación baja en la actividad {grade.activity.title}.',
                announcement_type='A',
                visibility_level='P',
                school=student.school
            )
            grade_announcement.students.set([student])
        return updated_grade

class QuarterGradeViewSet(ModelViewSet):

    queryset = models.QuarterGrade.objects.select_related('student', 'assignature', 'competence')
    serializer_class = serializers.QuarterGradeSerializer  
    permission_classes = [IsAuthenticated]

    # @action(detail=False, methods=['get'])
    # def forInstructor(self, request):
        # quarter = request.query_params.get('quarter') 
        # assignature = request.query_params.get('assignature')
        # quarter_grades = self.queryset.filter(quarter='Q1', assignature=7)
        # serializer = serializers.GetQuarterGradeForInstructorSerializer(quarter_grades, many=True)
        # return Response(serializer.data)

        # user = self.request.user
        # try:
        #     instructor = models.Instructor.objects.get(user_id=user.id)
        # except:
        #     return Response({"error": "Instructor not found for the current user"}, status=404)
        # quarter_grades = self.queryset.filter(assignature__instructor=instructor)
        # serializer = serializers.QuarterGradeSerializer(quarter_grades, many=True)
        # return Response(serializer.data)

class AnnouncementViewSet(ModelViewSet):

    queryset = models.Announcement.objects.select_related('created_by', 'school', 'assignature').prefetch_related('students', 'clases').order_by('-created_at')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateAnnouncementSerializer
        return serializers.GetAnnouncementSerializer
    
    @action(detail=False, methods=['get'])
    def byStudent(self, request):
        student = request.query_params.get('student')
        if not student:
            return Response({"error": "Student parameter is required"}, status=400)
        announcements = self.queryset.filter(student=student).order_by('-created_at')
        serializer = serializers.GetAnnouncementSerializer(announcements, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def byTutor(self, request):
        studentUid = request.query_params.get('student')
        date = request.query_params.get('date', now().date())
        try:    
            announcements = models.Announcement.objects.filter(students=studentUid, created_at__date=date).order_by('-created_at')
            if not announcements.exists():
                return Response([], status=200)
            serializer = serializers.GetAnnouncementSerializer(announcements, many=True)
            return Response(serializer.data)

        except models.Tutor.DoesNotExist:
            return Response({"error": "Tutor not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
    @action(detail=False, methods=['get'])
    def byAdmin(self, request):
        school = request.query_params.get('school')
        date_param = request.query_params.get('date', now().date()) 
        announcements = self.queryset.filter(school=school, created_at__date=date_param).order_by('-created_at')

        paginator = CustomPagination()
        try:    
            paginated_announcements = paginator.paginate_queryset(announcements, request)
        except:
            return Response([])

        serializer = serializers.GetAnnouncementSerializer(paginated_announcements, many=True)
        return Response(serializer.data) 

    @action(detail=False, methods=['get'])
    def byDate(self, request):
        """Retrieve announcements for the current day based on visibility level."""
        student_uid = request.query_params.get('student')
        date_param = request.query_params.get('date', now().date()) 
        try:
            student = models.Student.objects.get(uid=student_uid)
            print("student", student)
            # General announcements for the school
            general_announcements = self.queryset.filter(
                visibility_level='G',
                school=student.school,
                created_at__date=date_param
            )

            # Class announcements (for the student’s class)
            class_announcements = self.queryset.filter(
                visibility_level='C',
                clases=student.clase.id,
                created_at__date=date_param
            )

            # Assignature announcements (for subjects the student is in)
            
            assignature_announcements = self.queryset.filter(
                visibility_level='A',
                assignature__clase=student.clase,
                created_at__date=date_param
            )

            # Personal announcements (specifically assigned to the student)
            personal_announcements = self.queryset.filter(
                visibility_level='P',
                students=student,
                created_at__date=date_param
            )

            # Combine querysets
            # announcements = general_announcements | class_announcements | assignature_announcements | personal_announcements
            # announcements = announcements.distinct().order_by('-created_at')

            announcements = (general_announcements | class_announcements | assignature_announcements | personal_announcements).distinct().order_by('-created_at')


            serializer = serializers.GetAnnouncementSerializer(announcements, many=True)
            return Response(serializer.data)

        except models.Student.DoesNotExist:
            return Response({"error": "Student not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
    # def create(self, request, *args, **kwargs):

    #     student_id = request.data.get('student', []) 
    #     student = models.Student.objects.get(uid=student_id)

    #     try:
    #         tutor = models.Tutor.objects.get(students=student)
    #     except:
    #         return super().create(request, *args, **kwargs)

    #     tokens = FCMDevice.objects.filter(user=tutor.user)
    #     message = f'Tienes un nuevo mensaje sobre {student.first_name}'
    #     for token in tokens:
    #             send_push_notification(token.device_token, 'Nuevo Mensaje', message)
    #     return super().create(request, *args, **kwargs)

class HealthInfoViewSet(ModelViewSet):
    queryset = models.Health_Information.objects.select_related('student')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateHealthInfoSerializer
        return serializers.GetHealthInfoSerializer
    
class BirthInfoViewSet(ModelViewSet):
    queryset = models.Birth_Info.objects.select_related('student')
    permission_classes = [IsAuthenticated]


    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateBirthInfoSerializer
        return serializers.GetBirthInfoSerializer
    
class EmergencyContactViewSet(ModelViewSet):
    queryset = models.Emergency_Contact.objects.select_related('student')
    permission_classes = [IsAuthenticated]


    def get_serializer_class(self):
        if self.request.method == 'POST':
            return serializers.CreateEmergencyContactSerializer
        return serializers.GetEmergencyContactSerializer

class DeveloperViewSet(ModelViewSet):
    queryset = models.Developer.objects.select_related('user')
    serializer_class = serializers.DeveloperSerializer
    permission_classes = [IsAdminUser]

    @action(detail=False, methods=['get'])
    def me(self, request):
        user = self.request.user
        try:
            developer = self.queryset.get(user=user)
        except:
            return Response({"error": "Tutor not found for the current user"}, status=404)
        serializer = serializers.DeveloperSerializer(developer)
        return Response(serializer.data)
    
class TutorReadAgendaViewSet(ModelViewSet):

    queryset = models.TutorReadAgenda.objects.select_related('tutor', 'student')
    serializer_class = serializers.TutorReadAgendaSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def getReadAgenda(self, request):
        user_id = request.user.id
        student_id = request.query_params.get('student')
        date_param = request.query_params.get('date')

        if not student_id:
            return Response({"error": "Student ID is required"}, status=400)

        try:
            tutor = models.Tutor.objects.get(user_id=user_id)
        except models.Tutor.DoesNotExist:
            return Response({"error": "Tutor not found for the current user"}, status=404)

        try:
            student = models.Student.objects.get(uid=student_id)
        except models.Student.DoesNotExist:
            return Response({"error": "Student not found"}, status=404)

        if date_param is None:
            date_param = timezone.now().date()

        (read_agenda, created) = models.TutorReadAgenda.objects.get_or_create(
            tutor=tutor, student=student, agenda_date=date_param
        )

        serializer = serializers.TutorReadAgendaSerializer(read_agenda)
        return Response(serializer.data)
    
class TutorContactViewSet(ModelViewSet):

    queryset = models.TutorContact.objects.select_related('created_by', 'student')
    serializer_class = serializers.TutorContactSerializer
    permission_classes = [IsAuthenticated]

    # @action(detail=False, methods=['get'])
    # def getContactTutor(self, request):
    #     user = self.request.user
    #     student_uid = request.query_params.get('student')

    #     try:
    #         student = models.Student.objects.get(uid=student_uid)
    #     except models.Student.DoesNotExist:
    #         return Response({"error": "Student not found"}, status=404)
        
    #     (contact_user, created) = models.TutorContact.objects.get_or_create(
    #         student=student, contact_date=timezone.now().date(), created_by=user
    #     )

    #     serializer = serializers.TutorContactSerializer(contact_user)
    #     return Response(serializer.data)

class WhatsappMessageViewSet(ModelViewSet):
    queryset = models.WhatsappMessage.objects.select_related('created_by', 'student', 'school')
    serializer_class = serializers.WhatsappMessageSerialzer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        data = request.data
        student_uid = data.get('student')
        school_id = data.get('school')

        balance = models.Balance.objects.get(school_id=school_id)

        if not balance:
            return Response({"error": "Balance not found"}, status=400)
        if balance.amaount <= 0.1:
            return Response({"error": "Saldo Insuficiente"}, status=400)

        try:
            student = models.Student.objects.get(uid=student_uid)
            student_phone = student.tutor_phone  # Ensure this is in E.164 format (+1234567890)

            twilio_client = Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)

            message = twilio_client.messages.create(
                body=f"Hello {student.first_name}, this is a WhatsApp message from Django via Twilio!",
                from_="whatsapp:" + settings.TWILIO_PHONE_NUMBER,  # Ensure this is a WhatsApp-enabled number
                to="whatsapp:" + "+19085255111"  # Format properly for WhatsApp
            )

            response = super().create(request, *args, **kwargs)

            conversation_started = self.queryset.filter(
                student=student,
                created_at__date=timezone.now().date()
            ).exists()

            if conversation_started:
                balance.amaount -= 0.005
                balance.save()
            else:
                balance.amaount -= 0.025
                balance.save()


            return Response({
                "message_sid": message.sid,
                "status": message.status,
                "twilio_response": response.data
            }, status=status.HTTP_201_CREATED)

        except models.Student.DoesNotExist:
            return Response({"error": "Student not found"}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class TutorsAuthInfoViewSet(ModelViewSet):
    queryset = models.TutorAuthInfo.objects.select_related('student', 'school')
    serializer_class = serializers.TutorsAuthInfoSerializer
    permission_classes = [IsAdminUser]

    @action(detail=False, methods=['get'])
    def get_tutors_auth_info(self, request):

        # school= request.query_params.get('school')

        # if not school:
        #     return Response({"error": "School parameter is required"}, status=400)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"

        def get_grade_description(grade):
            grades_choices = {
                '1': '1ro',
                '2': '2do',
                '3': '3ro',
                '4': '4to',
                '5': '5to',
                '6': '6to',
            }
            return grades_choices.get(grade, grade)
        
        def get_section_description(section):
            if section == 'U':
                section = 'Unica'
            return f'{section}'
        
        def get_level_description(level):
            levels_choices = {
                'I': 'Inicial',
                'P': 'Primaria',
                'S': 'Secundaria',
            }
            return levels_choices.get(level, level)

        # Define border style
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                            top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="000066", end_color="000066", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        # Define headers

        headers = [
            "Usario",
            "Contraseña",
            "Grado",
            "Sección",
            "Nivel",
            "Nombres del Alunmno",
            "Apellidos del Alunmno",
            "Contacto",
        ]

        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        tutors = self.queryset.filter(school_id=1)

        for tutor in tutors:
            # Write data to the worksheet
            ws.append([
                tutor.username,
                tutor.password,
                get_grade_description(tutor.student.clase.grade),
                get_section_description(tutor.student.clase.section),
                get_level_description(tutor.student.clase.level),
                f'{tutor.student.first_name}' if tutor.student.first_name else '',
                f'{tutor.student.last_name}' if tutor.student.last_name else '',
                f'{tutor.student.tutor_phone}' if tutor.student.tutor_phone else 'Sin Contacto',
            ])
            # Apply border to the row
            # for cell in ws[-1]:
            #     cell.border = thin_border
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2


                # Prepare response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="data.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        
        return response

class BalanceViewSet(ModelViewSet):
    queryset = models.Balance.objects.select_related('school')
    serializer_class = serializers.BalanceSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='get_balance')
    def get_balance(self, request):
        school = request.query_params.get('school')
        if not school:
            return Response({"error": "School parameter is required"}, status=400)
        balance, created = self.queryset.get_or_create(school_id=school)
        serializer = serializers.BalanceSerializer(balance)
        return Response(serializer.data)
        
class LessonViewSet(ModelViewSet):
    queryset = models.Lesson.objects.select_related('assignature', 'instructor', 'classroom')
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user_id = self.request.user.id

        try:
            instructor = models.Instructor.objects.get(user_id=user_id)
        except models.Instructor.DoesNotExist:  
            return Response({"error": "Instructor not found for the current user"}, status=404)
        if self.request.user.is_superuser:
            return super().get_queryset()
        return self.queryset.filter(instructor=instructor).order_by('-created_at')
    
    def get_serializer_class(self):
        if self.request.method in SAFE_METHODS:
            return serializers.GetLessonSerializer
        return serializers.CreateLessonSerializer

    @action(detail=False, methods=['get'])
    def byAssignature(self, request):
        assignature = request.query_params.get('assignature')
        quarter = self.request.query_params.get('quarter')
        if not assignature:
            return Response({"error": "Assignature parameter is required"}, status=400)
        lessons = self.queryset.filter(assignature_id=assignature, quarter=quarter)
        serializer = serializers.GetLessonSerializer(lessons, many=True)
        return Response(serializer.data)

        
        







    
